
import io
import secrets
import json
import time
import os
import re
import traceback
import smtplib
import ssl
from datetime import datetime
from typing import Any, Dict, List
from urllib.parse import quote
from urllib.request import Request, urlopen

from flask import Flask, jsonify, make_response, request, send_file
from flask_cors import CORS
from email.mime.text import MIMEText
from openpyxl import Workbook
from pymongo import ASCENDING, DESCENDING, MongoClient, ReturnDocument
from pymongo.errors import DuplicateKeyError

def load_env_file(path: str = ".env"):
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, val = line.split("=", 1)
                key = key.strip()
                val = val.strip().strip('"').strip("'")
                if key and key not in os.environ:
                    os.environ[key] = val
    except FileNotFoundError:
        pass

load_env_file()
DEFAULT_SESSION = os.getenv("DEFAULT_SESSION", "2024_25")
CURRENT_SESSION = DEFAULT_SESSION
# MongoDB connection (hardcoded like exam backend)
MONGO_URI = "mongodb+srv://PSPS:2007@fee.4uslzr2.mongodb.net/?retryWrites=true&w=majority&appName=fee"
MONGO_DB_NAME = "school_fee"

SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")
ADMIN_LOGIN_ID = os.getenv("ADMIN_LOGIN_ID", "admin")
ADMIN_LOGIN_PASS = os.getenv("ADMIN_LOGIN_PASS", "admin123")
ADMIN_OTP_EMAIL = os.getenv("ADMIN_OTP_EMAIL", "").strip()

client = None
db = None
MONGO_INIT_ERROR = None
try:
    client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
    db = client[MONGO_DB_NAME]
    client.admin.command("ping")
except Exception as e:
    MONGO_INIT_ERROR = str(e)

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}}, supports_credentials=True)

upi_display_state = {"payload": None}

# simple in-memory OTP store: {login_id: {"otp": "123456", "expires": datetime}}
LOGIN_OTPS: Dict[str, Dict[str, Any]] = {}

MONTHS_ORDER = ["Annual", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar", "previousDue"]
EXAM_MONTHS_ORDER = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
DEFAULT_CLASSES = ["Nursery", "LKG", "UKG", "1st", "2nd", "3rd", "4th", "5th", "6th", "7th", "8th", "9th", "10th", "11th Arts", "12th Arts"]


def students_col():
    return db["students"]


def receipts_col():
    return db["receipts"]


def fees_col():
    return db["fee_structure"]

def fee_settings_col():
    return db["fee_settings"]

def sessions_col():
    return db["sessions"]


def counters_col():
    return db["counters"]


def exam_cfg_col():
    return db["exam_fee_config"]


def transport_routes_col():
    return db["transport_routes"]

def stationary_items_col():
    return db["stationary_items"]

def stationary_receipts_col():
    return db["stationary_receipts"]


def ensure_indexes():
    students_col().create_index([("session", ASCENDING), ("class_name", ASCENDING), ("roll", ASCENDING)], unique=True)
    students_col().create_index([("session", ASCENDING), ("id", ASCENDING)], unique=True)
    receipts_col().create_index([("session", ASCENDING), ("receipt_key", ASCENDING)], unique=True)
    receipts_col().create_index([("session", ASCENDING), ("id", ASCENDING)], unique=True)
    fees_col().create_index([("session", ASCENDING), ("class_name", ASCENDING)], unique=True)
    fee_settings_col().create_index([("session", ASCENDING)], unique=True)
    exam_cfg_col().create_index([("session", ASCENDING)], unique=True)
    transport_routes_col().create_index([("session", ASCENDING), ("route_name", ASCENDING)], unique=True)
    stationary_items_col().create_index([("session", ASCENDING), ("id", ASCENDING)], unique=True)
    stationary_items_col().create_index([("session", ASCENDING), ("name", ASCENDING)], unique=True)
    stationary_receipts_col().create_index([("session", ASCENDING), ("receipt_no", ASCENDING)], unique=True)
    stationary_receipts_col().create_index([("session", ASCENDING), ("class_name", ASCENDING), ("roll", ASCENDING)])
    sessions_col().create_index([("name", ASCENDING)], unique=True)


def get_next_sequence(key: str) -> int:
    doc = counters_col().find_one_and_update({"_id": key}, {"$inc": {"seq": 1}}, upsert=True, return_document=ReturnDocument.AFTER)
    return int(doc.get("seq", 1))


def sanitize_session_name(name: str) -> str:
    if not name:
        return CURRENT_SESSION
    safe = "".join(ch for ch in name if ch.isalnum() or ch in ("_", "-"))
    return safe or CURRENT_SESSION


def to_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except Exception:
        return default


def normalize_class_name(value: Any) -> str:
    return str(value or "").strip()


def fetch_master_students_for_session(session_name: str):
    """Fetch master students from students backend (best-effort)."""
    base = os.getenv("STUDENT_API_URL", "https://student-backend-117372286918.asia-south1.run.app/").strip()
    if not base:
        return []
    base = base.rstrip("/")
    url = f"{base}/students?session={quote(session_name)}"
    try:
        req = Request(url, headers={"Accept": "application/json"})
        with urlopen(req, timeout=8) as resp:
            raw = resp.read().decode("utf-8")
        payload = json.loads(raw)
        rows = payload if isinstance(payload, list) else (payload.get("students") if isinstance(payload, dict) else [])
        if not isinstance(rows, list):
            return []
        return rows
    except Exception:
        return []


def get_session_from_request() -> str:
    return sanitize_session_name(request.headers.get("X-Session") or request.args.get("session") or CURRENT_SESSION)


def get_previous_session_name(session_name: str):
    try:
        s, _ = session_name.split("_")
        start = int(s) - 1
        end = start + 1
        return f"{start}_{str(end)[2:]}"
    except Exception:
        return None


def default_month_structure():
    return {
        "Jan": {"status": "Due", "paid": 0, "due": 0}, "Feb": {"status": "Due", "paid": 0, "due": 0},
        "Mar": {"status": "Due", "paid": 0, "due": 0}, "Apr": {"status": "Due", "paid": 0, "due": 0},
        "May": {"status": "Due", "paid": 0, "due": 0}, "Jun": {"status": "Due", "paid": 0, "due": 0},
        "Jul": {"status": "Due", "paid": 0, "due": 0}, "Aug": {"status": "Due", "paid": 0, "due": 0},
        "Sep": {"status": "Due", "paid": 0, "due": 0}, "Oct": {"status": "Due", "paid": 0, "due": 0},
        "Nov": {"status": "Due", "paid": 0, "due": 0}, "Dec": {"status": "Due", "paid": 0, "due": 0},
        "Annual": {"status": "Due", "paid": 0, "due": 0},
    }


def ensure_months_normalized(months: Any):
    out = months if isinstance(months, dict) else {}
    for k, v in default_month_structure().items():
        if k not in out or not isinstance(out[k], dict):
            out[k] = v.copy()
            out[k]["exam_fee_applied"] = 0
        else:
            out[k]["status"] = out[k].get("status", "Due")
            out[k]["paid"] = int(out[k].get("paid", 0) or 0)
            out[k]["due"] = int(out[k].get("due", 0) or 0)
            out[k]["exam_fee_applied"] = int(out[k].get("exam_fee_applied", 0) or 0)
    return out


def sanitize_discount_config(raw: Any) -> Dict[str, Any]:
    cfg = raw if isinstance(raw, dict) else {}
    dtype = str(cfg.get("type", "fixed") or "fixed").strip().lower()
    if dtype not in {"fixed", "percentage"}:
        dtype = "fixed"
    scope = str(cfg.get("scope", "monthly") or "monthly").strip().lower()
    if scope not in {"monthly", "annual", "admission", "all"}:
        scope = "monthly"
    value = int(cfg.get("value", 0) or 0)
    if value < 0:
        value = 0
    note = str(cfg.get("note", "") or "").strip()[:300]
    return {
        "enabled": bool(cfg.get("enabled", False)),
        "type": dtype,
        "value": value,
        "scope": scope,
        "note": note,
    }


def normalize_exam_months(raw: Any) -> List[str]:
    if not isinstance(raw, list):
        return []
    out = []
    seen = set()
    for item in raw:
        m = str(item or "").strip()
        if m not in EXAM_MONTHS_ORDER:
            continue
        if m in seen:
            continue
        seen.add(m)
        out.append(m)
    out.sort(key=lambda x: EXAM_MONTHS_ORDER.index(x))
    return out


def normalize_exam_config(raw: Any) -> List[Dict[str, Any]]:
    if not isinstance(raw, list):
        return []
    out: List[Dict[str, Any]] = []
    for i, item in enumerate(raw):
        if not isinstance(item, dict):
            continue
        name = str(item.get("name", "") or "").strip()[:80] or f"Exam {i + 1}"
        months = normalize_exam_months(item.get("months", []))
        fees_in = item.get("class_fees", {})
        class_fees: Dict[str, int] = {}
        if isinstance(fees_in, dict):
            for k, v in fees_in.items():
                cls = normalize_class_name(k)
                if not cls:
                    continue
                class_fees[cls] = max(0, to_int(v, 0))
        out.append({"name": name, "months": months, "class_fees": class_fees})
    return out


def exam_target_for_class_month(exams: List[Dict[str, Any]], cls: str, month: str) -> int:
    total = 0
    cls_norm = normalize_class_name(cls)
    for ex in exams:
        months = ex.get("months", [])
        if month not in months:
            continue
        class_fees = ex.get("class_fees", {})
        total += max(0, to_int(class_fees.get(cls_norm, 0), 0))
    return total


def apply_exam_config_to_students(session_name: str, exams: List[Dict[str, Any]]):
    exams_norm = normalize_exam_config(exams)

    docs = list(students_col().find({"session": session_name}))
    for st in docs:
        cls = normalize_class_name(st.get("class_name"))
        months_obj = ensure_months_normalized(st.get("months") or {})

        changed = False
        for m in EXAM_MONTHS_ORDER:
            rec = months_obj.get(m)
            if not isinstance(rec, dict):
                continue
            target = exam_target_for_class_month(exams_norm, cls, m)
            current = int(rec.get("exam_fee_applied", 0) or 0)
            if target == current:
                continue
            delta = target - current
            rec["due"] = max(0, int(rec.get("due", 0) or 0) + delta)
            rec["exam_fee_applied"] = target
            paid = int(rec.get("paid", 0) or 0)
            due = int(rec.get("due", 0) or 0)
            rec["status"] = "Paid" if due == 0 else ("Partial" if paid > 0 else "Due")
            months_obj[m] = rec
            changed = True

        if changed:
            students_col().update_one(
                {"_id": st["_id"]},
                {"$set": {"months": months_obj, "updated_at": datetime.utcnow()}},
            )


def calc_carry_forward_amount(st: Dict[str, Any]) -> int:
    prev = int(st.get("previous_due", 0) or 0)
    months = ensure_months_normalized(st.get("months") or {})
    return prev + sum(int(rec.get("due", 0) or 0) for rec in months.values())


def send_bulk_email(to_list: List[str], subject: str, body: str):
    if not SMTP_USER or not SMTP_PASS:
        return False, "SMTP credentials not configured"
    if not to_list:
        return False, "No recipients"

    context = ssl.create_default_context()
    msg = MIMEText(body or "", "plain", "utf-8")
    msg["Subject"] = subject or "Fee Reminder"
    msg["From"] = SMTP_USER

    failed = []
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as server:
            server.starttls(context=context)
            server.login(SMTP_USER, SMTP_PASS)
            for email in to_list:
                msg["To"] = email
                try:
                    server.sendmail(SMTP_USER, [email], msg.as_string())
                except Exception:
                    failed.append(email)
    except Exception as e:
        return False, str(e)

    return True, failed


def ensure_session_exists(session_name: str):
    sname = sanitize_session_name(session_name)
    if sessions_col().find_one({"name": sname}, {"_id": 1}):
        return

    sessions_col().insert_one({"name": sname, "created_at": datetime.utcnow()})
    prev = get_previous_session_name(sname)
    if not prev:
        return

    prev_students = list(students_col().find({"session": prev}).sort("id", ASCENDING))
    for s in prev_students:
        roll = str(s.get("roll", ""))
        if students_col().find_one({"session": sname, "class_name": s.get("class_name"), "roll": roll}, {"_id": 1}):
            continue
        students_col().insert_one({
            "session": sname,
            "id": get_next_sequence(f"{sname}:student_id"),
            "name": s.get("name"),
            "father": s.get("father"),
            "class_name": s.get("class_name"),
            "roll": roll,
            "previous_due": calc_carry_forward_amount(s),
            "advance": 0,
            "months": ensure_months_normalized(default_month_structure()),
            "annual_charge": 0,
            "books_opt_in": bool(s.get("books_opt_in", False)),
            "scholarship": sanitize_discount_config(s.get("scholarship")),
            "concession": sanitize_discount_config(s.get("concession")),
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow(),
        })


def student_to_dict(d: Dict[str, Any]):
    months = ensure_months_normalized(d.get("months") or {})
    payment_methods = d.get("payment_methods") if isinstance(d.get("payment_methods"), dict) else {}
    return {
        "id": int(d.get("id", 0) or 0),
        "name": d.get("name"),
        "father": d.get("father"),
        "class_name": d.get("class_name"),
        "roll": str(d.get("roll", "")),
        "admission_no": str(d.get("admission_no") or d.get("admission") or ""),
        "previous_due": int(d.get("previous_due", 0) or 0),
        "advance": int(d.get("advance", 0) or 0),
        "uses_transport": bool(d.get("uses_transport", False)),
        "transport_route": str(d.get("transport_route", "") or ""),
        "transport_fee": int(d.get("transport_fee", 0) or 0),
        "transport_bus_no": str(d.get("transport_bus_no", "") or ""),
        "transport_months": int(d.get("transport_months", 0) or 0),
        "months": months,
        "annual_charge": int(d.get("annual_charge", months.get("Annual", {}).get("paid", 0)) or 0),
        "books_opt_in": bool(d.get("books_opt_in", False)),
        "last_payment_method": str(d.get("last_payment_method", "") or ""),
        "payment_methods": {
            "cash": int(payment_methods.get("cash", 0) or 0),
            "upi": int(payment_methods.get("upi", 0) or 0),
            "bank": int(payment_methods.get("bank", 0) or 0),
        },
        "scholarship": sanitize_discount_config(d.get("scholarship")),
        "concession": sanitize_discount_config(d.get("concession")),
    }


def receipt_to_dict(d: Dict[str, Any]):
    return {
        "id": int(d.get("id", 0) or 0),
        "session": d.get("session"),
        "name": d.get("name"),
        "father": d.get("father"),
        "payment_type": d.get("payment_type", ""),
        "admission_no": d.get("admission_no", ""),
        "new_admission": bool(d.get("new_admission", False)),
        "admission_charge": int(d.get("admission_charge", 0) or 0),
        "registration_fee": int(d.get("registration_fee", 0) or 0),
        "id_card_fee": int(d.get("id_card_fee", 0) or 0),
        "books_charge": int(d.get("books_charge", 0) or 0),
        "books_opt_in": bool(d.get("books_opt_in", False)),
        "scholarship": sanitize_discount_config(d.get("scholarship")),
        "concession": sanitize_discount_config(d.get("concession")),
        "class_name": d.get("class_name"),
        "roll": str(d.get("roll", "")),
        "date": d.get("date"),
        "total_paid": int(d.get("total_paid", 0) or 0),
        "total_due": int(d.get("total_due", 0) or 0),
        "advance": int(d.get("advance", 0) or 0),
        "latest_paid": int(d.get("latest_paid", 0) or 0),
        "annual_charge": int(d.get("annual_charge", 0) or 0),
        "receipt_number": d.get("receipt_number"),
        "receipt_key": d.get("receipt_key"),
        "months": d.get("months") or {},
    }


def apply_payment_to_student_months_and_prev(months: Dict[str, Any], prev_due: int, payment: int):
    months = ensure_months_normalized(months or {})
    remaining = int(payment or 0)
    pay_prev = min(remaining, int(prev_due or 0))
    prev_due = int(prev_due or 0) - pay_prev
    remaining -= pay_prev

    for m in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Annual"]:
        if remaining <= 0:
            break
        rec = months.get(m) or {"paid": 0, "due": 0, "status": "Due"}
        due_amt = int(rec.get("due", 0) or 0)
        if due_amt <= 0:
            continue
        pay = min(remaining, due_amt)
        rec["paid"] = int(rec.get("paid", 0) or 0) + pay
        rec["due"] = due_amt - pay
        rec["status"] = "Paid" if rec["due"] <= 0 else "Partial"
        months[m] = rec
        remaining -= pay
    return months, prev_due, remaining


@app.before_request
def before_request_switch_db():
    if request.method == "OPTIONS":
        resp = make_response(jsonify({"ok": True, "reason": "preflight"}), 200)
        resp.headers["Access-Control-Allow-Origin"] = request.headers.get("Origin", "*")
        resp.headers["Access-Control-Allow-Methods"] = "GET,POST,PUT,DELETE,OPTIONS"
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type,X-Session"
        return resp

    try:
        if db is None:
            raise RuntimeError(MONGO_INIT_ERROR or "MongoDB is not initialized. Check MONGO_URI.")
        ensure_indexes()
        ensure_session_exists(get_session_from_request())
    except Exception as e:
        resp = make_response(jsonify({"success": False, "error": f"MongoDB connection failed: {str(e)}"}), 500)
        resp.headers["Access-Control-Allow-Origin"] = request.headers.get("Origin", "*")
        return resp


@app.route("/")
def home():
    return jsonify({"success": True, "message": "Backend running", "session": get_session_from_request(), "db": "mongodb"})


@app.route("/health")
def health():
    return jsonify({"status": "ok", "session": get_session_from_request(), "db": "mongodb"})


@app.route("/email/send-bulk", methods=["POST"])
def email_send_bulk():
    data = request.get_json() or {}
    emails = data.get("emails")
    subject = str(data.get("subject") or "Fee Reminder")
    message = str(data.get("message") or "").strip()
    if not isinstance(emails, list) or not emails:
        return jsonify({"success": False, "message": "Missing email list"}), 400
    if not message:
        return jsonify({"success": False, "message": "Message is empty"}), 400

    ok, result = send_bulk_email([e for e in emails if isinstance(e, str) and e.strip()], subject, message)
    if not ok:
        return jsonify({"success": False, "message": result}), 400

    failed = result if isinstance(result, list) else []
    return jsonify({
        "success": True,
        "sent": len(emails) - len(failed),
        "failed": failed
    })


@app.route("/auth/otp/request", methods=["POST"])
def auth_otp_request():
    data = request.get_json() or {}
    login_id = str(data.get("login_id") or "").strip()
    password = str(data.get("password") or "").strip()
    if not login_id or not password:
        return jsonify({"success": False, "message": "Missing login_id or password"}), 400

    if login_id != ADMIN_LOGIN_ID or password != ADMIN_LOGIN_PASS:
        return jsonify({"success": False, "message": "Invalid credentials"}), 401

    if not SMTP_USER or not SMTP_PASS:
        return jsonify({"success": False, "message": "SMTP credentials not configured"}), 400

    otp = str(secrets.randbelow(1000000)).zfill(6)
    expires = datetime.utcnow().timestamp() + 5 * 60  # 5 minutes
    LOGIN_OTPS[login_id] = {"otp": otp, "expires": expires}

    to_email = ADMIN_OTP_EMAIL or SMTP_USER
    ok, err = send_bulk_email([to_email], "OTP for Fee Login", f"Your OTP is {otp}. It expires in 5 minutes.")
    if not ok:
        return jsonify({"success": False, "message": err}), 400

    return jsonify({"success": True, "message": "OTP sent"})


@app.route("/auth/otp/verify", methods=["POST"])
def auth_otp_verify():
    data = request.get_json() or {}
    login_id = str(data.get("login_id") or "").strip()
    otp = str(data.get("otp") or "").strip()
    if not login_id or not otp:
        return jsonify({"success": False, "message": "Missing login_id or otp"}), 400

    rec = LOGIN_OTPS.get(login_id)
    if not rec:
        return jsonify({"success": False, "message": "OTP not requested"}), 400
    if datetime.utcnow().timestamp() > float(rec.get("expires", 0)):
        LOGIN_OTPS.pop(login_id, None)
        return jsonify({"success": False, "message": "OTP expired"}), 400
    if otp != rec.get("otp"):
        return jsonify({"success": False, "message": "Invalid OTP"}), 401

    LOGIN_OTPS.pop(login_id, None)
    return jsonify({"success": True, "token": f"fee_admin_{login_id}_{int(datetime.utcnow().timestamp())}"})


@app.route("/debug/list_dbs")
def debug_list_dbs():
    sessions = [d.get("name") for d in sessions_col().find({}, {"name": 1, "_id": 0}) if d.get("name")]
    return jsonify(sorted(sessions))


@app.route("/session/list")
def session_list():
    sessions = set(d.get("name") for d in sessions_col().find({}, {"name": 1, "_id": 0}) if d.get("name"))
    sessions.add(DEFAULT_SESSION)

    def sort_key(val: str):
        m = re.match(r"^(\d+)_", val or "")
        return int(m.group(1)) if m else 0

    return jsonify({"success": True, "sessions": sorted(sessions, key=sort_key)})


@app.route("/session/create_auto", methods=["POST"])
def create_auto_session():
    data = request.get_json() or {}
    from_session = sanitize_session_name(data.get("from_session"))
    extra_fee = int(data.get("extra_fee", 0) or 0)

    if not from_session:
        return jsonify({"success": False, "message": "Missing from_session"}), 400

    try:
        s, e = from_session.split("_")
        new_session = f"{int(s) + 1}_{str(int(e) + 1)[-2:]}"
    except Exception:
        return jsonify({"success": False, "message": "Invalid session format"}), 400

    ensure_session_exists(new_session)
    if extra_fee:
        students_col().update_many({"session": new_session}, {"$inc": {"previous_due": extra_fee}, "$set": {"updated_at": datetime.utcnow()}})

    return jsonify({"success": True, "new_session": new_session})


@app.route("/delete_session", methods=["POST"])
def delete_session():
    try:
        sname = sanitize_session_name(request.headers.get("X-Session"))
        students_col().delete_many({"session": sname})
        receipts_col().delete_many({"session": sname})
        fees_col().delete_many({"session": sname})
        sessions_col().delete_one({"name": sname})
        counters_col().delete_many({"_id": {"$regex": f"^{re.escape(sname)}:"}})
        return jsonify({"success": True, "message": "Session deleted successfully"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/student/add", methods=["POST"])
def add_student():
    data = request.get_json() or {}
    name = data.get("name")
    cls = data.get("class_name")
    roll = str(data.get("roll")) if data.get("roll") is not None else None
    if not name or not cls or roll is None:
        return jsonify({"success": False, "message": "Missing name/class/roll"}), 400

    sname = get_session_from_request()
    months = ensure_months_normalized(data.get("months", {}))
    if students_col().find_one({"session": sname, "class_name": cls, "roll": roll}, {"_id": 1}):
        return jsonify({"success": False, "message": "Student exists"}), 200

    doc = {
        "session": sname,
        "id": get_next_sequence(f"{sname}:student_id"),
        "name": name,
        "father": data.get("father"),
        "class_name": cls,
        "roll": roll,
        "admission_no": str(data.get("admission_no") or data.get("admission") or "").strip(),
        "previous_due": int(data.get("previous_due", 0) or 0),
        "advance": int(data.get("advance", 0) or 0),
        "uses_transport": bool(data.get("uses_transport", False)),
        "transport_route": str(data.get("transport_route", "") or "").strip(),
        "transport_fee": int(data.get("transport_fee", 0) or 0),
        "transport_bus_no": str(data.get("transport_bus_no", "") or "").strip(),
        "transport_months": max(0, min(12, int(data.get("transport_months", 0) or 0))),
        "months": months,
        "last_payment_method": str(data.get("last_payment_method", "") or ""),
        "payment_methods": {
            "cash": int(((data.get("payment_methods") or {}).get("cash", 0)) or 0),
            "upi": int(((data.get("payment_methods") or {}).get("upi", 0)) or 0),
            "bank": int(((data.get("payment_methods") or {}).get("bank", 0)) or 0),
        },
        "annual_charge": int(months.get("Annual", {}).get("paid", 0) or 0),
        "books_opt_in": bool(data.get("books_opt_in", False)),
        "scholarship": sanitize_discount_config(data.get("scholarship")),
        "concession": sanitize_discount_config(data.get("concession")),
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow(),
    }
    try:
        students_col().insert_one(doc)
    except DuplicateKeyError:
        return jsonify({"success": False, "message": "Student exists"}), 200

    return jsonify({"success": True, "student": student_to_dict(doc)})


@app.route("/students")
def get_students():
    sname = get_session_from_request()
    class_name = (request.args.get("class_name") or request.args.get("class") or "").strip()
    q = {"session": sname}
    if class_name:
        q["class_name"] = class_name
    docs = students_col().find(q, {"_id": 0}).sort([("class_name", ASCENDING), ("roll", ASCENDING)])
    return jsonify({"success": True, "students": [student_to_dict(d) for d in docs]})


@app.route("/student/<class_name>/<roll>")
def get_single_student(class_name, roll):
    sname = get_session_from_request()
    doc = students_col().find_one({"session": sname, "class_name": class_name, "roll": str(roll)}, {"_id": 0})
    if not doc:
        return jsonify({"success": False, "message": "Not found"}), 404
    return jsonify({"success": True, "student": student_to_dict(doc)})

@app.route("/update_student", methods=["POST", "OPTIONS"])
def update_student():
    if request.method == "OPTIONS":
        return jsonify({"status": "ok"})

    data = request.json or {}
    class_name = data.get("class")
    roll = str(data.get("roll")) if data.get("roll") is not None else None
    if not class_name or roll is None:
        return jsonify({"success": False, "message": "Missing class or roll"}), 400

    sname = get_session_from_request()
    sdata = data.get("student", {})
    student = students_col().find_one({"session": sname, "class_name": class_name, "roll": roll})
    if not student:
        return jsonify({"success": False, "message": "Student not found"}), 404

    months = ensure_months_normalized(sdata.get("months", student.get("months", {})))
    pm_in = sdata.get("payment_methods", student.get("payment_methods", {}))
    if not isinstance(pm_in, dict):
        pm_in = {}
    payment_methods = {
        "cash": int(pm_in.get("cash", ((student.get("payment_methods") or {}).get("cash", 0)) or 0)),
        "upi": int(pm_in.get("upi", ((student.get("payment_methods") or {}).get("upi", 0)) or 0)),
        "bank": int(pm_in.get("bank", ((student.get("payment_methods") or {}).get("bank", 0)) or 0)),
    }
    new_class = sdata.get("class_name", student.get("class_name"))
    new_roll = str(sdata.get("roll", student.get("roll")))

    if (new_class != class_name or new_roll != roll) and students_col().find_one({"session": sname, "class_name": new_class, "roll": new_roll, "id": {"$ne": student.get("id")}}, {"_id": 1}):
        return jsonify({"success": False, "message": "Student exists with new class/roll"}), 409

    students_col().update_one(
        {"session": sname, "id": student.get("id")},
        {
            "$set": {
                "name": sdata.get("name", student.get("name")),
                "father": sdata.get("father", student.get("father")),
                "admission_no": str(sdata.get("admission_no", student.get("admission_no", "")) or sdata.get("admission") or student.get("admission") or "").strip(),
                "previous_due": int(sdata.get("previous_due", student.get("previous_due", 0)) or 0),
                "advance": int(sdata.get("advance", student.get("advance", 0)) or 0),
                "uses_transport": bool(sdata.get("uses_transport", student.get("uses_transport", False))),
                "transport_route": str(sdata.get("transport_route", student.get("transport_route", "")) or "").strip(),
                "transport_fee": int(sdata.get("transport_fee", student.get("transport_fee", 0)) or 0),
                "transport_bus_no": str(sdata.get("transport_bus_no", student.get("transport_bus_no", "")) or "").strip(),
                "transport_months": max(0, min(12, int(sdata.get("transport_months", student.get("transport_months", 0)) or 0))),
                "months": months,
                "last_payment_method": str(sdata.get("last_payment_method", student.get("last_payment_method", "")) or ""),
                "payment_methods": payment_methods,
                "annual_charge": int(months.get("Annual", {}).get("paid", 0) or 0),
                "books_opt_in": bool(sdata.get("books_opt_in", student.get("books_opt_in", False))),
                "scholarship": sanitize_discount_config(sdata.get("scholarship", student.get("scholarship"))),
                "concession": sanitize_discount_config(sdata.get("concession", student.get("concession"))),
                "class_name": new_class,
                "roll": new_roll,
                "updated_at": datetime.utcnow(),
            }
        },
    )
    return jsonify({"success": True, "message": "Updated"})


@app.route("/student/delete", methods=["POST"])
def delete_student():
    data = request.json or {}
    cls = data.get("class")
    roll = str(data.get("roll")) if data.get("roll") is not None else None
    if not cls or roll is None:
        return jsonify({"success": False, "message": "Missing class or roll"}), 400

    sname = get_session_from_request()
    if not students_col().find_one({"session": sname, "class_name": cls, "roll": roll}, {"_id": 1}):
        return jsonify({"success": False, "message": "Not found"}), 404

    receipts_col().delete_many({"session": sname, "class_name": cls, "roll": roll})
    students_col().delete_one({"session": sname, "class_name": cls, "roll": roll})
    return jsonify({"success": True, "message": "Deleted"})


@app.route("/receipt/add", methods=["POST"])
def add_receipt():
    try:
        data = request.json or {}
        required = ["name", "father", "class", "roll", "date", "totalPaid", "totalDue", "advance", "months", "receiptKey"]
        for key in required:
            if key not in data:
                return jsonify({"success": False, "message": f"Missing {key}"}), 400

        sname = get_session_from_request()
        months_raw = data.get("months", {})
        months = {}

        if isinstance(months_raw, list):
            for item in months_raw:
                key = item.get("month") or item.get("name")
                if not key:
                    continue
                months[key] = {
                    "paid": int(item.get("paid", 0) or 0),
                    "due": int(item.get("due", 0) or 0),
                    "status": item.get("status", ""),
                    "purpose": item.get("purpose", ""),
                    "extra": int(item.get("extra", 0) or 0),
                    "date": item.get("date", ""),
                    "exam_fee_applied": int(item.get("exam_fee_applied", 0) or 0),
                    "transport_fee_applied": int(item.get("transport_fee_applied", 0) or 0),
                }
        elif isinstance(months_raw, dict):
            for k, v in months_raw.items():
                vv = v if isinstance(v, dict) else {}
                months[k] = {
                    "paid": int(vv.get("paid", 0) or 0),
                    "due": int(vv.get("due", 0) or 0),
                    "status": vv.get("status", ""),
                    "purpose": vv.get("purpose", ""),
                    "extra": int(vv.get("extra", 0) or 0),
                    "date": vv.get("date", ""),
                    "exam_fee_applied": int(vv.get("exam_fee_applied", 0) or 0),
                    "transport_fee_applied": int(vv.get("transport_fee_applied", 0) or 0),
                }

        old = receipts_col().find_one({"session": sname, "receipt_key": data["receiptKey"]}, {"_id": 0, "receipt_number": 1})
        if old:
            return jsonify({"success": True, "message": "Duplicate ignored", "receipt_number": old.get("receipt_number")})

        def receipt_signature(payload_months):
            keys = sorted(payload_months.keys(), key=lambda x: str(x))
            clean = {}
            for k in keys:
                v = payload_months.get(k) or {}
                clean[str(k)] = {
                    "paid": int(v.get("paid", 0) or 0),
                    "due": int(v.get("due", 0) or 0),
                    "status": str(v.get("status", "") or ""),
                    "purpose": str(v.get("purpose", "") or ""),
                    "extra": int(v.get("extra", 0) or 0),
                    "date": str(v.get("date", "") or ""),
                    "exam_fee_applied": int(v.get("exam_fee_applied", 0) or 0),
                    "transport_fee_applied": int(v.get("transport_fee_applied", 0) or 0),
                }
            base = {
                "class": str(data.get("class") or ""),
                "roll": str(data.get("roll") or ""),
                "name": str(data.get("name") or ""),
                "father": str(data.get("father") or ""),
                "total_paid": int(data.get("totalPaid") or 0),
                "total_due": int(data.get("totalDue") or 0),
                "advance": int(data.get("advance") or 0),
                "months": clean,
            }
            return json.dumps(base, sort_keys=True, separators=(",", ":"))

        sig = receipt_signature(months)
        last = receipts_col().find_one(
            {"session": sname, "class_name": data["class"], "roll": str(data["roll"])},
            {"_id": 0, "receipt_number": 1, "receipt_signature": 1, "months": 1, "total_paid": 1, "total_due": 1, "advance": 1},
            sort=[("id", DESCENDING)],
        )
        if last:
            last_sig = last.get("receipt_signature") or receipt_signature(last.get("months") or {})
            if last_sig == sig:
                return jsonify({"success": True, "message": "Duplicate ignored", "receipt_number": last.get("receipt_number")})

        rid = get_next_sequence(f"{sname}:receipt_id")
        receipt_number = f"{sname}-{data['class']}-{data['roll']}-{rid:06d}"
        annual_paid = int(months.get("Annual", {}).get("paid", 0) or 0)

        receipts_col().insert_one({
            "session": sname,
            "id": rid,
            "student_id": None,
            "name": data["name"],
            "father": data["father"],
            "payment_type": str(data.get("paid_type", data.get("payment_type", "")) or "").lower(),
            "admission_no": str(data.get("admission_no", "") or ""),
            "new_admission": bool(data.get("new_admission", False)),
            "admission_charge": int(data.get("admission_charge", 0) or 0),
            "registration_fee": int(data.get("registration_fee", 0) or 0),
            "id_card_fee": int(data.get("id_card_fee", 0) or 0),
            "books_charge": int(data.get("books_charge", 0) or 0),
            "books_opt_in": bool(data.get("books_opt_in", False)),
            "scholarship": sanitize_discount_config(data.get("scholarship")),
            "concession": sanitize_discount_config(data.get("concession")),
            "class_name": data["class"],
            "roll": str(data["roll"]),
            "date": data["date"],
            "total_paid": int(data["totalPaid"]),
            "total_due": int(data["totalDue"]),
            "advance": int(data["advance"]),
            "latest_paid": int(data.get("latest_paid", 0) or 0),
            "months": months,
            "receipt_key": data["receiptKey"],
            "receipt_signature": sig,
            "annual_charge": annual_paid,
            "receipt_number": receipt_number,
            "created_at": datetime.utcnow(),
        })

        # Do not re-apply payment to student ledger here.
        # Student months are already updated in student_details page (submit/mark paid).
        # Re-applying on receipt save causes duplicate month payments after generating receipt.

        return jsonify({"success": True, "message": "Receipt saved", "receipt_number": receipt_number})
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/receipt/history")
def receipt_history():
    sname = get_session_from_request()
    cls = (request.args.get("class_name") or request.args.get("class") or "").strip()
    roll = str(request.args.get("roll") or "").strip()
    # Backward-compatible behavior:
    # - If page/page_size are provided, return paginated history.
    # - Otherwise return full history (old behavior).
    page_arg = request.args.get("page")
    size_arg = request.args.get("page_size")

    if page_arg is None and size_arg is None:
        q = {"session": sname}
        if cls:
            q["class_name"] = cls
        if roll:
            q["roll"] = roll
        docs = receipts_col().find(q, {"_id": 0}).sort("id", DESCENDING)
        return jsonify({"success": True, "history": [receipt_to_dict(d) for d in docs]})

    try:
        page = max(1, int(page_arg or 1))
    except Exception:
        page = 1
    try:
        page_size = int(size_arg or 50)
    except Exception:
        page_size = 50
    page_size = min(max(page_size, 1), 200)

    q = {"session": sname}
    if cls:
        q["class_name"] = cls
    if roll:
        q["roll"] = roll
    total = receipts_col().count_documents(q)
    total_pages = max(1, (total + page_size - 1) // page_size)
    if page > total_pages:
        page = total_pages

    skip = (page - 1) * page_size
    docs = (
        receipts_col()
        .find(q, {"_id": 0})
        .sort("id", DESCENDING)
        .skip(skip)
        .limit(page_size)
    )
    return jsonify({
        "success": True,
        "history": [receipt_to_dict(d) for d in docs],
        "page": page,
        "page_size": page_size,
        "total": int(total),
        "total_pages": int(total_pages),
    })


@app.route("/receipt/delete/<int:id>", methods=["DELETE"])
def delete_receipt(id):
    try:
        sname = get_session_from_request()
        deleted = receipts_col().delete_one({"session": sname, "id": int(id)}).deleted_count
        if not deleted:
            return jsonify({"success": False, "message": "Receipt not found"}), 404
        return jsonify({"success": True, "message": "Receipt deleted"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/receipt/delete_all", methods=["DELETE"])
def delete_all_receipts():
    try:
        receipts_col().delete_many({"session": get_session_from_request()})
        return jsonify({"success": True, "message": "All deleted"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/upi/display", methods=["GET", "POST", "OPTIONS"])
def upi_display():
    if request.method == "OPTIONS":
        return jsonify({"status": "ok"})
    if request.method == "POST":
        data = request.get_json() or {}
        # Allow posting an arbitrary payload (for richer tablet screens like result/attendance/fee-structure).
        # Keep backward compatibility with the older POST shape used for QR.
        posted_payload = data.get("payload") if isinstance(data.get("payload"), dict) else None
        if posted_payload is not None:
            if not posted_payload.get("ts"):
                posted_payload["ts"] = int(time.time() * 1000)
            if not posted_payload.get("expiresSec"):
                posted_payload["expiresSec"] = int(data.get("expiresSec", 600) or 600)
            upi_display_state["payload"] = posted_payload
            return jsonify({"success": True})

        upi_display_state["payload"] = {
            "amount": data.get("amount"),
            "upiUrl": data.get("upiUrl"),
            "expiresSec": int(data.get("expiresSec", 600) or 600),
            "ts": int(data.get("ts", time.time() * 1000) or int(time.time() * 1000)),
            "studentName": data.get("studentName"),
            "mode": data.get("mode", "qr"),
            "student": data.get("student"),
        }
        return jsonify({"success": True})

    # Allow GET with query params to set payload (helps avoid CORS/preflight issues)
    if request.args.get("clear") == "1":
        upi_display_state["payload"] = None
        return jsonify({"success": True})

    if request.args.get("mode") == "student":
        upi_display_state["payload"] = {
            "mode": "student",
            "student": {
                "name": request.args.get("name"),
                "father": request.args.get("father"),
                "class": request.args.get("class"),
                "roll": request.args.get("roll"),
                "admission_no": request.args.get("admission_no"),
                "photoUrl": request.args.get("photoUrl"),
                "totalDue": request.args.get("totalDue"),
                "latestPaid": request.args.get("latestPaid"),
                "status": request.args.get("status"),
            },
            "ts": int(request.args.get("ts", time.time() * 1000) or int(time.time() * 1000)),
        }
        return jsonify({"success": True})

    if request.args.get("mode") == "thanks":
        upi_display_state["payload"] = {
            "mode": "thanks",
            "ts": int(request.args.get("ts", time.time() * 1000) or int(time.time() * 1000)),
        }
        return jsonify({"success": True})

    if request.args.get("mode") == "receipt":
        upi_display_state["payload"] = {
            "mode": "receipt",
            "receipt": {
                "name": request.args.get("name"),
                "class": request.args.get("class"),
                "roll": request.args.get("roll"),
                "admission_no": request.args.get("admission_no"),
                "totalPaid": request.args.get("totalPaid"),
                "totalDue": request.args.get("totalDue"),
                "latestPaid": request.args.get("latestPaid"),
                "date": request.args.get("date"),
                "paymentMode": request.args.get("paymentMode"),
                "status": request.args.get("status"),
            },
            "ts": int(request.args.get("ts", time.time() * 1000) or int(time.time() * 1000)),
        }
        return jsonify({"success": True})

    if request.args.get("mode") == "cash":
        upi_display_state["payload"] = {
            "mode": "cash",
            "cash": {
                "name": request.args.get("name"),
                "admission_no": request.args.get("admission_no"),
                "amount": request.args.get("amount"),
                "totalPaid": request.args.get("totalPaid"),
            },
            "ts": int(request.args.get("ts", time.time() * 1000) or int(time.time() * 1000)),
        }
        return jsonify({"success": True})

    if request.args.get("amount") and request.args.get("upiUrl"):
        upi_display_state["payload"] = {
            "amount": request.args.get("amount"),
            "upiUrl": request.args.get("upiUrl"),
            "expiresSec": int(request.args.get("expiresSec", 600) or 600),
            "ts": int(request.args.get("ts", time.time() * 1000) or int(time.time() * 1000)),
            "studentName": request.args.get("studentName"),
            "mode": request.args.get("mode", "qr"),
        }
        return jsonify({"success": True})

    payload = upi_display_state.get("payload")
    if not payload:
        return jsonify({"success": True, "payload": None})
    expires_ms = int(payload.get("expiresSec", 600) or 600) * 1000
    ts = int(payload.get("ts", 0) or 0)
    if ts and (int(time.time() * 1000) - ts) > expires_ms:
        upi_display_state["payload"] = None
        return jsonify({"success": True, "payload": None})
    return jsonify({"success": True, "payload": payload})


def stationary_item_to_dict(d: Dict[str, Any]):
    return {
        "id": int(d.get("id", 0) or 0),
        "name": d.get("name"),
        "price": int(d.get("price", 0) or 0),
    }


@app.route("/stationary/items", methods=["GET", "POST"])
def stationary_items():
    sname = get_session_from_request()
    if request.method == "GET":
        rows = list(stationary_items_col().find({"session": sname}, {"_id": 0}).sort("name", ASCENDING))
        return jsonify({"success": True, "items": [stationary_item_to_dict(r) for r in rows]})

    data = request.get_json() or {}
    name = str(data.get("name") or "").strip()
    price = max(0, to_int(data.get("price"), 0))
    if not name:
        return jsonify({"success": False, "message": "Missing item name"}), 400

    existing = stationary_items_col().find_one({"session": sname, "name": name})
    if existing:
        stationary_items_col().update_one(
            {"session": sname, "name": name},
            {"$set": {"price": price, "updated_at": datetime.utcnow()}},
        )
        updated = stationary_items_col().find_one({"session": sname, "name": name}, {"_id": 0})
        return jsonify({"success": True, "item": stationary_item_to_dict(updated), "updated": True})

    item_id = get_next_sequence(f"{sname}:stationary_item_id")
    doc = {"session": sname, "id": item_id, "name": name, "price": price, "created_at": datetime.utcnow()}
    stationary_items_col().insert_one(doc)
    return jsonify({"success": True, "item": stationary_item_to_dict(doc), "created": True})


@app.route("/stationary/items/<int:item_id>", methods=["PUT", "DELETE"])
def stationary_item_update_delete(item_id: int):
    sname = get_session_from_request()
    if request.method == "DELETE":
        deleted = stationary_items_col().delete_one({"session": sname, "id": int(item_id)}).deleted_count
        return jsonify({"success": True, "deleted": int(deleted)})

    data = request.get_json() or {}
    name = str(data.get("name") or "").strip()
    price = max(0, to_int(data.get("price"), 0))
    if not name:
        return jsonify({"success": False, "message": "Missing item name"}), 400

    if stationary_items_col().find_one({"session": sname, "name": name, "id": {"$ne": int(item_id)}}, {"_id": 1}):
        return jsonify({"success": False, "message": "Item name already exists"}), 409

    stationary_items_col().update_one(
        {"session": sname, "id": int(item_id)},
        {"$set": {"name": name, "price": price, "updated_at": datetime.utcnow()}},
    )
    updated = stationary_items_col().find_one({"session": sname, "id": int(item_id)}, {"_id": 0})
    if not updated:
        return jsonify({"success": False, "message": "Item not found"}), 404
    return jsonify({"success": True, "item": stationary_item_to_dict(updated)})


@app.route("/stationary/receipt", methods=["POST"])
def stationary_receipt_add():
    data = request.get_json() or {}
    sname = get_session_from_request()
    required = ["student", "items"]
    for key in required:
        if key not in data:
            return jsonify({"success": False, "message": f"Missing {key}"}), 400

    student = data.get("student") or {}
    name = str(student.get("name") or "").strip()
    father = str(student.get("father") or "").strip()
    class_name = str(student.get("class_name") or student.get("class") or "").strip()
    roll = str(student.get("roll") or "").strip()
    admission_no = str(student.get("admission_no") or student.get("admission") or "").strip()
    if not name or not class_name or not roll:
        return jsonify({"success": False, "message": "Missing student details"}), 400

    items_in = data.get("items")
    if not isinstance(items_in, list) or not items_in:
        return jsonify({"success": False, "message": "No items selected"}), 400

    items_out = []
    total = 0
    for raw in items_in:
        if not isinstance(raw, dict):
            continue
        item_id = int(raw.get("id", 0) or 0)
        name_in = str(raw.get("name") or "").strip()
        qty = max(1, to_int(raw.get("qty"), 1))
        price = max(0, to_int(raw.get("price"), 0))

        if item_id:
            db_item = stationary_items_col().find_one({"session": sname, "id": item_id}, {"_id": 0})
            if db_item:
                name_in = db_item.get("name") or name_in
                price = int(db_item.get("price", price) or price)

        if not name_in:
            continue

        line_total = price * qty
        total += line_total
        items_out.append({
            "id": item_id,
            "name": name_in,
            "price": price,
            "qty": qty,
            "total": line_total,
        })

    if not items_out:
        return jsonify({"success": False, "message": "No valid items"}), 400

    receipt_no = get_next_sequence(f"{sname}:stationary_receipt")
    receipt = {
        "session": sname,
        "receipt_no": int(receipt_no),
        "date": data.get("date") or datetime.utcnow().strftime("%Y-%m-%d"),
        "student_name": name,
        "father": father,
        "class_name": class_name,
        "roll": roll,
        "admission_no": admission_no,
        "items": items_out,
        "total": int(total),
        "added_to_fee": bool(data.get("added_to_fee", False)),
        "payment_mode": str(data.get("payment_mode") or "").strip().lower(),
        "payment_status": str(data.get("payment_status") or "unpaid").strip().lower(),
        "paid_at": str(data.get("paid_at") or "").strip(),
        "status_updated_at": "",
        "created_at": datetime.utcnow(),
    }
    stationary_receipts_col().insert_one(receipt)
    # Remove non-JSON types before returning
    receipt.pop("_id", None)
    receipt.pop("created_at", None)
    return jsonify({"success": True, "receipt": receipt})


@app.route("/stationary/add-to-fee", methods=["POST"])
def stationary_add_to_fee():
    data = request.get_json() or {}
    sname = get_session_from_request()
    student = data.get("student") or {}
    items_in = data.get("items") or []
    date_str = str(data.get("date") or "").strip()

    class_name = str(student.get("class_name") or student.get("class") or "").strip()
    roll = str(student.get("roll") or "").strip()
    if not class_name or not roll:
        return jsonify({"success": False, "message": "Missing class or roll"}), 400

    if not isinstance(items_in, list) or not items_in:
        return jsonify({"success": False, "message": "No items selected"}), 400

    # Calculate total and build purpose text
    total = 0
    parts = []
    for raw in items_in:
        if not isinstance(raw, dict):
            continue
        name_in = str(raw.get("name") or "").strip()
        qty = max(1, to_int(raw.get("qty"), 1))
        price = max(0, to_int(raw.get("price"), 0))
        if not name_in:
            continue
        total += price * qty
        parts.append(f"{name_in}x{qty}")

    if total <= 0:
        return jsonify({"success": False, "message": "Invalid total"}), 400

    # Month key from date or today
    try:
        if date_str:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
        else:
            dt = datetime.utcnow()
    except Exception:
        dt = datetime.utcnow()
    month_key = dt.strftime("%b")  # Jan, Feb, ...

    student_doc = students_col().find_one({"session": sname, "class_name": class_name, "roll": roll})
    if not student_doc:
        return jsonify({"success": False, "message": "Student not found"}), 404

    months = ensure_months_normalized(student_doc.get("months") or {})
    rec = months.get(month_key) or {"paid": 0, "due": 0, "status": "Due"}
    rec["due"] = int(rec.get("due", 0) or 0) + int(total)
    rec["date"] = date_str or dt.strftime("%Y-%m-%d")
    prev_purpose = str(rec.get("purpose") or "").strip()
    purpose = ""
    new_items = ", ".join(parts)
    base_label = "Stationary: "
    if prev_purpose:
        if "Stationary:" in prev_purpose:
            # Merge items under a single Stationary label.
            other_parts = []
            existing_items = []
            for seg in prev_purpose.split("|"):
                seg = seg.strip()
                if not seg:
                    continue
                if seg.startswith(base_label):
                    existing_items.append(seg.replace(base_label, "").strip())
                else:
                    other_parts.append(seg)
            merged_items = ", ".join([x for x in existing_items if x] + ([new_items] if new_items else []))
            stationary_part = f"{base_label}{merged_items}".strip()
            combined = " | ".join([p for p in other_parts + [stationary_part] if p])
            rec["purpose"] = combined
            purpose = stationary_part
        else:
            purpose = f"{base_label}{new_items}".strip()
            rec["purpose"] = f"{prev_purpose} | {purpose}".strip(" |")
    else:
        rec["purpose"] = f"{base_label}{new_items}".strip()
        purpose = rec["purpose"]
    paid = int(rec.get("paid", 0) or 0)
    due = int(rec.get("due", 0) or 0)
    rec["status"] = "Paid" if due == 0 else ("Partial" if paid > 0 else "Due")
    months[month_key] = rec

    students_col().update_one(
        {"session": sname, "class_name": class_name, "roll": roll},
        {"$set": {"months": months, "updated_at": datetime.utcnow()}},
    )

    return jsonify({
        "success": True,
        "message": "Stationary added to fee ledger",
        "month": month_key,
        "total": int(total),
        "purpose": purpose
    })


@app.route("/stationary/receipt/<int:receipt_no>")
def stationary_receipt_get(receipt_no: int):
    sname = get_session_from_request()
    doc = stationary_receipts_col().find_one({"session": sname, "receipt_no": int(receipt_no)}, {"_id": 0})
    if not doc:
        return jsonify({"success": False, "message": "Not found"}), 404
    return jsonify({"success": True, "receipt": doc})


@app.route("/stationary/receipt/<int:receipt_no>/payment", methods=["PUT"])
def stationary_receipt_payment_update(receipt_no: int):
    sname = get_session_from_request()
    data = request.get_json() or {}

    payment_mode = str(data.get("payment_mode") or "").strip().lower()
    payment_status = str(data.get("payment_status") or "").strip().lower()
    paid_at = str(data.get("paid_at") or "").strip()

    allowed_modes = {"", "cash", "upi"}
    allowed_statuses = {"unpaid", "pending", "paid", "not_paid", "cancelled"}

    if payment_mode not in allowed_modes:
        return jsonify({"success": False, "message": "Invalid payment mode"}), 400
    if payment_status not in allowed_statuses:
        return jsonify({"success": False, "message": "Invalid payment status"}), 400

    update_doc = {
        "payment_mode": payment_mode,
        "payment_status": payment_status,
        "status_updated_at": datetime.utcnow().isoformat(),
    }
    if payment_status == "paid":
        update_doc["paid_at"] = paid_at or datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    elif "paid_at" in data:
        update_doc["paid_at"] = paid_at

    stationary_receipts_col().update_one(
        {"session": sname, "receipt_no": int(receipt_no)},
        {"$set": update_doc},
    )
    updated = stationary_receipts_col().find_one({"session": sname, "receipt_no": int(receipt_no)}, {"_id": 0})
    if not updated:
        return jsonify({"success": False, "message": "Not found"}), 404
    return jsonify({"success": True, "receipt": updated})


@app.route("/stationary/receipt/<int:receipt_no>", methods=["DELETE"])
def stationary_receipt_delete(receipt_no: int):
    sname = get_session_from_request()
    deleted = stationary_receipts_col().delete_one({"session": sname, "receipt_no": int(receipt_no)}).deleted_count
    if not deleted:
        return jsonify({"success": False, "message": "Not found"}), 404
    return jsonify({"success": True, "message": "Deleted"})


@app.route("/stationary/receipts")
def stationary_receipts_list():
    sname = get_session_from_request()
    cls = (request.args.get("class_name") or request.args.get("class") or "").strip()
    roll = str(request.args.get("roll") or "").strip()
    added_to_fee = request.args.get("added_to_fee")
    page = to_int(request.args.get("page", 1), 1)
    page_size = to_int(request.args.get("page_size", 10), 10)
    if page < 1:
        page = 1
    if page_size < 1:
        page_size = 10
    if page_size > 100:
        page_size = 100
    q = {"session": sname}
    if cls:
        q["class_name"] = cls
    if roll:
        q["roll"] = roll
    if added_to_fee is not None:
        q["added_to_fee"] = str(added_to_fee).lower() in ("1","true","yes","y","on")
    total = stationary_receipts_col().count_documents(q)
    docs = list(
        stationary_receipts_col()
        .find(q, {"_id": 0})
        .sort("receipt_no", DESCENDING)
        .skip((page - 1) * page_size)
        .limit(page_size)
    )
    total_pages = (total + page_size - 1) // page_size if page_size else 1
    return jsonify({
        "success": True,
        "receipts": docs,
        "page": page,
        "page_size": page_size,
        "total": int(total),
        "total_pages": int(total_pages),
    })

@app.route("/fees/get")
def fees_get():
    sname = get_session_from_request()
    existing = list(fees_col().find({"session": sname}, {"_id": 0}))
    existing_classes = {normalize_class_name(d.get("class_name")) for d in existing if normalize_class_name(d.get("class_name"))}
    missing = [c for c in DEFAULT_CLASSES if c not in existing_classes]

    if missing:
        fees_col().insert_many([
            {
                "session": sname,
                "class_name": c,
                "monthly_fee": 0,
                "annual_charge": 0,
                "admission_charge": 0,
                "books_charge": 0,
                "exam_charge": 0,
                "created_at": datetime.utcnow(),
                "updated_at": datetime.utcnow(),
            }
            for c in missing
        ])

    # Backfill old records that used "admission_fee" or had no admission field.
    # This keeps old DB data consistent with the current API.
    old_rows = list(
        fees_col().find(
            {"session": sname, "admission_charge": {"$exists": False}},
            {"_id": 1, "admission_fee": 1},
        )
    )
    for row in old_rows:
        mapped = to_int(row.get("admission_fee", 0), 0)
        fees_col().update_one(
            {"_id": row["_id"]},
            {"$set": {"admission_charge": mapped, "admission_fee": mapped, "updated_at": datetime.utcnow()}},
        )

    old_exam_rows = list(
        fees_col().find(
            {"session": sname, "exam_charge": {"$exists": False}},
            {"_id": 1, "exam_fee": 1},
        )
    )
    for row in old_exam_rows:
        mapped = to_int(row.get("exam_fee", 0), 0)
        fees_col().update_one(
            {"_id": row["_id"]},
            {"$set": {"exam_charge": mapped, "exam_fee": mapped, "updated_at": datetime.utcnow()}},
        )

    docs = list(fees_col().find({"session": sname}, {"_id": 0}))
    merged_by_class: Dict[str, Dict[str, Any]] = {}
    for d in docs:
        cls = normalize_class_name(d.get("class_name"))
        if not cls:
            continue
        current = merged_by_class.get(cls)
        if current is None:
            merged_by_class[cls] = d
            continue

        cur_time = current.get("updated_at") or current.get("created_at") or datetime.min
        new_time = d.get("updated_at") or d.get("created_at") or datetime.min
        if new_time >= cur_time:
            merged_by_class[cls] = d

    normalized_docs = list(merged_by_class.values())
    normalized_docs.sort(key=lambda d: (DEFAULT_CLASSES.index(normalize_class_name(d.get("class_name"))) if normalize_class_name(d.get("class_name")) in DEFAULT_CLASSES else 999, normalize_class_name(d.get("class_name"))))
    fees = []
    for i, d in enumerate(normalized_docs):
        fees.append({
            "id": i + 1,
            "class_name": normalize_class_name(d.get("class_name")),
            "monthly_fee": to_int(d.get("monthly_fee", 0), 0),
            "annual_charge": to_int(d.get("annual_charge", 0), 0),
            "admission_charge": to_int(d.get("admission_charge", d.get("admission_fee", 0)), 0),
            "books_charge": to_int(d.get("books_charge", 0), 0),
            "exam_charge": to_int(d.get("exam_charge", d.get("exam_fee", 0)), 0),
        })
    return jsonify({"success": True, "fees": fees})

@app.route("/fees/settings", methods=["GET"])
def fees_settings_get():
    sname = get_session_from_request()
    doc = fee_settings_col().find_one({"session": sname}, {"_id": 0}) or {}
    return jsonify({
        "success": True,
        "registration_fee": to_int(doc.get("registration_fee", 0), 0),
        "id_card_fee": to_int(doc.get("id_card_fee", 0), 0),
    })

@app.route("/fees/settings", methods=["POST"])
def fees_settings_set():
    data = request.json or {}
    sname = get_session_from_request()
    registration_fee = to_int(data.get("registration_fee", 0), 0)
    id_card_fee = to_int(data.get("id_card_fee", 0), 0)
    fee_settings_col().update_one(
        {"session": sname},
        {"$set": {
            "registration_fee": registration_fee,
            "id_card_fee": id_card_fee,
            "updated_at": datetime.utcnow(),
        }, "$setOnInsert": {"created_at": datetime.utcnow()}},
        upsert=True,
    )
    return jsonify({"success": True, "registration_fee": registration_fee, "id_card_fee": id_card_fee})


@app.route("/fees/update", methods=["POST"])
def update_fee():
    data = request.json or {}
    cls = normalize_class_name(data.get("class_name"))
    if not cls:
        return jsonify({"success": False, "message": "Missing class_name"}), 400

    admission_value = data.get("admission_charge", data.get("admission_fee", 0))
    exam_value = data.get("exam_charge", data.get("exam_fee", 0))
    sname = get_session_from_request()
    fees_col().update_one(
        {"session": sname, "class_name": cls},
        {
            "$set": {
                "monthly_fee": to_int(data.get("monthly_fee", 0), 0),
                "annual_charge": to_int(data.get("annual_charge", 0), 0),
                "admission_charge": to_int(admission_value, 0),
                "admission_fee": to_int(admission_value, 0),
                "books_charge": to_int(data.get("books_charge", 0), 0),
                "exam_charge": to_int(exam_value, 0),
                "exam_fee": to_int(exam_value, 0),
                "updated_at": datetime.utcnow(),
            },
            "$setOnInsert": {"created_at": datetime.utcnow()},
        },
        upsert=True,
    )
    saved = fees_col().find_one({"session": sname, "class_name": cls}, {"_id": 0})
    return jsonify({
        "success": True,
        "message": "Updated",
        "fee": {
            "class_name": cls,
            "monthly_fee": to_int((saved or {}).get("monthly_fee", 0), 0),
            "annual_charge": to_int((saved or {}).get("annual_charge", 0), 0),
            "admission_charge": to_int((saved or {}).get("admission_charge", (saved or {}).get("admission_fee", 0)), 0),
            "books_charge": to_int((saved or {}).get("books_charge", 0), 0),
            "exam_charge": to_int((saved or {}).get("exam_charge", (saved or {}).get("exam_fee", 0)), 0),
        },
    })


@app.route("/fees/update_many", methods=["POST"])
def update_fees_many():
    data = request.json or {}
    rows = data.get("fees")
    if not isinstance(rows, list) or not rows:
        return jsonify({"success": False, "message": "Missing fees list"}), 400

    sname = get_session_from_request()
    updated = 0
    for row in rows:
        if not isinstance(row, dict):
            continue
        cls = normalize_class_name(row.get("class_name"))
        if not cls:
            continue

        admission_value = row.get("admission_charge", row.get("admission_fee", 0))
        exam_value = row.get("exam_charge", row.get("exam_fee", 0))
        fees_col().update_one(
            {"session": sname, "class_name": cls},
            {
                "$set": {
                    "monthly_fee": to_int(row.get("monthly_fee", 0), 0),
                    "annual_charge": to_int(row.get("annual_charge", 0), 0),
                    "admission_charge": to_int(admission_value, 0),
                    "admission_fee": to_int(admission_value, 0),
                    "books_charge": to_int(row.get("books_charge", 0), 0),
                    "exam_charge": to_int(exam_value, 0),
                    "exam_fee": to_int(exam_value, 0),
                    "updated_at": datetime.utcnow(),
                },
                "$setOnInsert": {"created_at": datetime.utcnow()},
            },
            upsert=True,
        )
        updated += 1

    return jsonify({"success": True, "message": f"Updated {updated} classes", "updated": updated})


@app.route("/fees/setup_defaults")
def setup_fees():
    defaults = {
        "Nursery": 1200, "LKG": 1300, "UKG": 1300,
        "1st": 1300, "2nd": 1300, "3rd": 1300, "4th": 1400, "5th": 1400,
        "6th": 1500, "7th": 1500, "8th": 1700, "9th": 1900, "10th": 1900,
        "11th_Medical": 2200, "11th_Commerce": 2100, "11th_Art": 2100,
        "12th_Medical": 2200, "12th_Commerce": 2100, "12th_Art": 2100,
    }
    sname = get_session_from_request()
    for cls, fee in defaults.items():
        fees_col().update_one(
            {"session": sname, "class_name": cls},
            {
                "$set": {"monthly_fee": int(fee), "updated_at": datetime.utcnow()},
                "$setOnInsert": {"annual_charge": 0, "admission_charge": 0, "exam_charge": 0, "exam_fee": 0, "created_at": datetime.utcnow()},
            },
            upsert=True,
        )
    return jsonify({"success": True, "message": "Inserted"})


@app.route("/exam/config", methods=["GET", "POST"])
def exam_config():
    sname = get_session_from_request()
    if request.method == "GET":
        doc = exam_cfg_col().find_one({"session": sname}, {"_id": 0})
        exams = normalize_exam_config((doc or {}).get("exams", []))
        return jsonify({"success": True, "session": sname, "exams": exams})

    data = request.json or {}
    exams = normalize_exam_config(data.get("exams", []))
    exam_cfg_col().update_one(
        {"session": sname},
        {"$set": {"session": sname, "exams": exams, "updated_at": datetime.utcnow()}, "$setOnInsert": {"created_at": datetime.utcnow()}},
        upsert=True,
    )
    apply_exam_config_to_students(sname, exams)
    return jsonify({"success": True, "message": "Exam config updated", "session": sname, "exams": exams})


@app.route("/transport/routes", methods=["GET", "POST", "DELETE"])
def transport_routes():
    sname = get_session_from_request()

    if request.method == "GET":
        docs = list(
            transport_routes_col()
            .find({"session": sname}, {"_id": 0})
            .sort([("route_name", ASCENDING)])
        )
        routes = [{
            "route_name": normalize_class_name(d.get("route_name")),
            "route_fee": to_int(d.get("route_fee", 0), 0),
            "bus_no": str(d.get("bus_no", "") or "").strip(),
        } for d in docs if normalize_class_name(d.get("route_name"))]
        return jsonify({"success": True, "session": sname, "routes": routes})

    if request.method == "POST":
        data = request.get_json() or {}
        route_name = normalize_class_name(data.get("route_name"))
        route_fee = max(0, to_int(data.get("route_fee", 0), 0))
        bus_no = str(data.get("bus_no", "") or "").strip()
        if not route_name:
            return jsonify({"success": False, "message": "Missing route_name"}), 400

        transport_routes_col().update_one(
            {"session": sname, "route_name": route_name},
            {
                "$set": {
                    "route_fee": route_fee,
                    "bus_no": bus_no,
                    "updated_at": datetime.utcnow(),
                },
                "$setOnInsert": {
                    "session": sname,
                    "route_name": route_name,
                    "created_at": datetime.utcnow(),
                },
            },
            upsert=True,
        )
        return jsonify({"success": True, "message": "Route saved"})

    data = request.get_json() or {}
    route_name = normalize_class_name(data.get("route_name"))
    if not route_name:
        return jsonify({"success": False, "message": "Missing route_name"}), 400
    deleted = transport_routes_col().delete_one({"session": sname, "route_name": route_name}).deleted_count
    return jsonify({"success": True, "deleted": int(deleted)})


@app.route("/transport/students", methods=["GET"])
def transport_students():
    sname = get_session_from_request()
    class_name = normalize_class_name(request.args.get("class_name"))

    master_rows = fetch_master_students_for_session(sname)
    fee_rows = list(students_col().find({"session": sname}, {"_id": 0}))
    fee_map = {
        f"{normalize_class_name(x.get('class_name'))}|{str(x.get('roll', '')).strip()}": x
        for x in fee_rows
    }

    out = []
    for m in master_rows:
        cls = normalize_class_name(m.get("class_name") or m.get("class"))
        if not cls:
            continue
        if class_name and cls != class_name:
            continue
        roll = str(m.get("rollno") or m.get("roll") or "").strip()
        if not roll:
            continue
        key = f"{cls}|{roll}"
        fs = fee_map.get(key, {})

        uses_transport = bool(fs.get("uses_transport", False))
        out.append({
            "name": str(m.get("student_name") or m.get("name") or fs.get("name") or "").strip(),
            "class_name": cls,
            "roll": roll,
            "admission_no": str(m.get("admission_no") or fs.get("admission_no") or "").strip(),
            "parent_mobile": str(m.get("parent_mobile") or m.get("mobile") or fs.get("parent_mobile") or "").strip(),
            "uses_transport": uses_transport,
            "transport_route": str(fs.get("transport_route", "") or "").strip(),
            "transport_fee": to_int(fs.get("transport_fee", 0), 0),
            "transport_bus_no": str(fs.get("transport_bus_no", "") or "").strip(),
            "transport_months": max(0, min(12, to_int(fs.get("transport_months", 12), 12))),
        })

    out.sort(key=lambda x: (x.get("class_name", ""), to_int(x.get("roll", 0), 0)))
    return jsonify({"success": True, "students": out})


@app.route("/transport/students/save", methods=["POST"])
def transport_students_save():
    data = request.get_json() or {}
    sname = get_session_from_request()
    rows = data.get("students")
    if not isinstance(rows, list):
        return jsonify({"success": False, "message": "Missing students list"}), 400

    saved = 0
    for row in rows:
        if not isinstance(row, dict):
            continue
        cls = normalize_class_name(row.get("class_name"))
        roll = str(row.get("roll", "") or "").strip()
        if not cls or not roll:
            continue

        uses_transport = bool(row.get("uses_transport", False))
        route = normalize_class_name(row.get("transport_route"))
        fee = max(0, to_int(row.get("transport_fee", 0), 0))
        bus_no = str(row.get("transport_bus_no", "") or "").strip()
        months = max(1, min(12, to_int(row.get("transport_months", 12), 12)))

        # Ensure fee-student row exists; keep minimal fields if creating new.
        existing = students_col().find_one({"session": sname, "class_name": cls, "roll": roll}, {"_id": 1, "name": 1, "father": 1})
        if not existing:
            students_col().insert_one({
                "session": sname,
                "id": get_next_sequence(f"{sname}:student_id"),
                "name": str(row.get("name", "") or "").strip(),
                "father": "",
                "class_name": cls,
                "roll": roll,
                "previous_due": 0,
                "advance": 0,
                "months": ensure_months_normalized(default_month_structure()),
                "annual_charge": 0,
                "scholarship": sanitize_discount_config({}),
                "concession": sanitize_discount_config({}),
                "created_at": datetime.utcnow(),
                "updated_at": datetime.utcnow(),
            })

        students_col().update_one(
            {"session": sname, "class_name": cls, "roll": roll},
            {
                "$set": {
                    "uses_transport": uses_transport,
                    "transport_route": route if uses_transport else "",
                    "transport_fee": fee if uses_transport else 0,
                    "transport_bus_no": bus_no if uses_transport else "",
                    "transport_months": months if uses_transport else 0,
                    "updated_at": datetime.utcnow(),
                }
            },
        )
        saved += 1

    return jsonify({"success": True, "saved": saved, "message": f"Saved {saved} students"})


@app.route("/getClassList/<session_name>")
def get_class_list(session_name):
    sname = sanitize_session_name(session_name)
    classes = students_col().distinct("class_name", {"session": sname})
    return jsonify([{"class_name": c} for c in sorted([x for x in classes if x])])


@app.route("/backup/save_all_to_drive", methods=["POST"])
def backup_all_sessions_to_drive():
    sessions = [d.get("name") for d in sessions_col().find({}, {"name": 1, "_id": 0}) if d.get("name")]
    files = [{"session": s, "success": False, "message": "MongoDB mode enabled. Use MongoDB Atlas backup/snapshots.", "file_id": None} for s in sessions]
    return jsonify({"success": True, "message": "Processed", "files": files})


@app.route("/export/excel")
def export_excel():
    sname = get_session_from_request()
    students = list(students_col().find({"session": sname}, {"_id": 0}).sort("id", ASCENDING))
    receipts = list(receipts_col().find({"session": sname}, {"_id": 0}).sort("id", ASCENDING))

    def normalize_export_class(value: Any) -> str:
        text = str(value or "").strip()
        text = re.sub(r"\s+", " ", text)
        # Keep sectioned classes mapped to core class fee (e.g. "11th A" -> "11th").
        text = re.sub(r"\s+[AB]$", "", text, flags=re.IGNORECASE).strip()
        return text

    # Include master students as well, so export is not limited to fee-ledger-created rows.
    # This keeps Excel aligned with index view where students can come from master backend.
    master_rows = fetch_master_students_for_session(sname)
    existing_keys = {
        f"{normalize_export_class(s.get('class_name'))}|{str(s.get('roll', '')).strip()}": 1
        for s in students
    }
    next_id = max([int(s.get("id", 0) or 0) for s in students] + [0]) + 1
    for m in master_rows:
        cls = normalize_export_class(m.get("class_name") or m.get("class") or "")
        roll = str(m.get("rollno") or m.get("roll") or "").strip()
        if not cls or not roll:
            continue
        key = f"{cls}|{roll}"
        if key in existing_keys:
            continue
        students.append({
            "session": sname,
            "id": next_id,
            "name": m.get("student_name") or m.get("name") or "",
            "father": m.get("father_name") or m.get("father") or "",
            "class_name": cls,
            "roll": roll,
            "previous_due": 0,
            "advance": 0,
            "months": ensure_months_normalized(default_month_structure()),
            "annual_charge": 0,
        })
        existing_keys[key] = 1
        next_id += 1

    students.sort(key=lambda s: (normalize_export_class(s.get("class_name")), str(s.get("roll", ""))))

    # Fee defaults by class (used when student fee ledger/month rows are missing).
    fee_docs = list(fees_col().find({"session": sname}, {"_id": 0}))
    fee_by_class: Dict[str, Dict[str, int]] = {}
    for d in fee_docs:
        cls = normalize_export_class(d.get("class_name"))
        if not cls:
            continue
        fee_by_class[cls] = {
            "monthly_fee": to_int(d.get("monthly_fee", 0), 0),
            "annual_charge": to_int(d.get("annual_charge", 0), 0),
            "admission_charge": to_int(d.get("admission_charge", d.get("admission_fee", 0)), 0),
        }

    def get_class_defaults(cls_name: str, new_admission: bool):
        fd = fee_by_class.get(normalize_export_class(cls_name), {})
        monthly = int(fd.get("monthly_fee", 0) or 0)
        annual = int(fd.get("annual_charge", 0) or 0)
        admission = int(fd.get("admission_charge", 0) or 0) if new_admission else 0
        return monthly, annual + admission

    def norm_month_record(rec: Any, due_default: int):
        # Match index behavior: if month record is missing, treat as Due by default fee.
        if not isinstance(rec, dict):
            return 0, int(due_default), ("Paid" if int(due_default) <= 0 else "Due")
        paid = to_int(rec.get("paid", 0), 0)
        due = to_int(rec.get("due", due_default), int(due_default))
        # If record exists but untouched (paid=0,due=0), apply class default due.
        # This is common for students that exist in master backend but were never opened in fee ledger.
        if paid == 0 and due == 0 and int(due_default) > 0:
            due = int(due_default)
        status = str(rec.get("status", "") or "").strip()
        if status not in ("Paid", "Partial", "Due"):
            if due <= 0 and paid > 0:
                status = "Paid"
            elif paid > 0 and due > 0:
                status = "Partial"
            else:
                status = "Due"
        return paid, due, status

    def build_student_row(st: Dict[str, Any]):
        cls_name = normalize_export_class(st.get("class_name"))
        roll = str(st.get("roll", ""))
        prev_due_default = to_int(st.get("previous_due", 0), 0)
        new_adm = bool(st.get("new_admission", False))
        monthly_default, annual_default = get_class_defaults(cls_name, new_adm)
        months = st.get("months") if isinstance(st.get("months"), dict) else {}

        row = [
            int(st.get("id", 0) or 0),
            st.get("name"),
            st.get("father"),
            cls_name,
            roll,
            prev_due_default,
            to_int(st.get("advance", 0), 0),
        ]

        total_paid, total_due = 0, 0
        for m in MONTHS_ORDER:
            if m == "previousDue":
                due_default = prev_due_default
            elif m == "Annual":
                due_default = annual_default
            else:
                due_default = monthly_default
            paid, due, status = norm_month_record(months.get(m), due_default)
            row.extend([paid, status])
            total_paid += paid
            total_due += due

        row.extend([total_paid, total_due])
        return row, total_paid, total_due

    wb = Workbook()
    ws = wb.active
    ws.title = "Students Fee Report"

    header = ["ID", "Name", "Father", "Class", "Roll", "Previous Due", "Advance"]
    for m in MONTHS_ORDER:
        header.extend([f"{m} Paid", f"{m} Status"])
    header.extend(["Total Paid", "Total Due"])
    ws.append(header)

    total_school_paid, total_school_due, class_summary = 0, 0, {}
    for st in students:
        row, total_paid, total_due = build_student_row(st)
        ws.append(row)

        total_school_paid += total_paid
        total_school_due += total_due
        cls = st.get("class_name")
        if cls not in class_summary:
            class_summary[cls] = {"students": 0, "paid": 0, "due": 0}
        class_summary[cls]["students"] += 1
        class_summary[cls]["paid"] += total_paid
        class_summary[cls]["due"] += total_due

    # Add one sheet per class containing only that class's students.
    def safe_sheet_name(name: str, used_names: set):
        base = str(name or "").strip() or "Unknown"
        base = re.sub(r"[\\/*?:\[\]]", "_", base)[:31] or "Unknown"
        candidate = base
        i = 1
        while candidate in used_names:
            suffix = f"_{i}"
            candidate = (base[: 31 - len(suffix)] + suffix)[:31]
            i += 1
        used_names.add(candidate)
        return candidate

    students_by_class: Dict[str, list] = {}
    for st in students:
        cls = str(st.get("class_name") or "Unknown").strip() or "Unknown"
        students_by_class.setdefault(cls, []).append(st)

    used_sheet_names = {"Students Fee Report"}
    for cls in sorted(students_by_class.keys(), key=lambda x: x.lower()):
        ws_cls = wb.create_sheet(safe_sheet_name(cls, used_sheet_names))
        ws_cls.append(header)
        for st in students_by_class[cls]:
            row, _, _ = build_student_row(st)
            ws_cls.append(row)

    ws2 = wb.create_sheet("Summary")
    ws2.append(["SUMMARY REPORT"])
    ws2.append([""])
    ws2.append(["Total Students", len(students)])
    ws2.append(["Total Collection (Paid)", total_school_paid])
    ws2.append(["Total Due", total_school_due])
    ws2.append([""])
    ws2.append(["Class", "Students", "Total Paid", "Total Due"])
    for cls, val in class_summary.items():
        ws2.append([cls, val["students"], val["paid"], val["due"]])

    ws3 = wb.create_sheet("Receipts")
    ws3.append(["ID", "Name", "Father", "Class", "Roll", "Date", "Total Paid", "Total Due", "Advance", "Annual Charge", "Receipt Number", "Months (JSON)"])
    for r in receipts:
        ws3.append([
            int(r.get("id", 0) or 0), r.get("name"), r.get("father"), r.get("class_name"), r.get("roll"), r.get("date"),
            int(r.get("total_paid", 0) or 0), int(r.get("total_due", 0) or 0), int(r.get("advance", 0) or 0),
            int(r.get("annual_charge", 0) or 0), r.get("receipt_number"), json.dumps(r.get("months") or {}),
        ])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, download_name=f"School_Report_{sname}.xlsx", as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")), debug=False)

